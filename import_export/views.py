from django.shortcuts import render, redirect
from django.contrib import messages

def import_students_view(request):
    from main.models import Student, Grade


    if request.method == 'POST':
        
        excel_file = request.FILES['excel_file']
        # print(excel_file)

        import openpyxl
        from main.models import Student  # Student Model from main app

     
        workbook = openpyxl.load_workbook(excel_file)

        # Select the active worksheet
        sheet = workbook.active

        # Read the data
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row
            # Split the class name and section while handling potential non-string types
            row0 = str(row[0]).strip()
            if isinstance(row0, str) and '-' in row0:
                class_name, section = map(str.strip, row0.split('-'))
            else:
                class_name = str(row0).strip()
                section = ''

            data.append({
                'grade': class_name,
                'section': section,
                'roll_no': row[1],
                'full_name': row[2]
            })


        #group data by class and section if possible
        from itertools import groupby
        data = groupby(data, key=lambda x: (x['grade'], x['section']))


        for key, values in data:
            grade_name, section = key
            
            try:
                if section != '':
                    grade = Grade.objects.get(name=grade_name, section=section)
                else:
                    grade = Grade.objects.get(name=grade_name)
            except Grade.DoesNotExist:
                grade = Grade.objects.create(name=grade_name, section=section, teacher=None)

            print(grade, grade_name)
            for value in values:
                # print(value)
                try:
                    student = Student.objects.get(grade=grade, roll_no=value['roll_no'], name=value['full_name'])
                except Student.DoesNotExist:
                    Student.objects.create(
                        grade=grade,  
                        roll_no=value['roll_no'],
                        name=value['full_name']
                    )

       

        message = "Students imported successfully"
        messages.success(request, message)
        return redirect('import_students_view')
    


    context = {
        'students': Student.objects.all()
    }
    
    return render(request, 'import_export/students.html', context)






def export_student_data_to_excel_view(request):
    import openpyxl
    from main.models import Student  # Student Model from main app
    from django.http import HttpResponse


    # Create a new Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students Data'
    
    # Add headers to the worksheet
    headers = ['Class', 'Roll No', 'Full Name'] 
    ws.append(headers)
    
    # Query the data from the database
    student_data = Student.objects.all().values_list('grade__name', 'grade__section', 'roll_no', 'name')


    #group students by class and section if possible
    temp_dict = {}
    for data in student_data:
        if data[1] != '':
            grade = f"{data[0]} - {data[1]}"
        else:
            grade = data[0]

        if grade not in temp_dict:
            temp_dict[grade] = [(data[2], data[3])]
        else:
            temp_dict[grade].append((data[2], data[3]))
        
    

    # Write student_data to the worksheet
    for key, values in temp_dict.items():
        grade = key
        for roll_no, name in values:
            new_row = (grade, roll_no, name)
            ws.append(new_row)
            # print(new_row)
    
    # Prepare the response as a downloadable Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    import datetime
    response['Content-Disposition'] = f'attachment; filename=students_data-{datetime.datetime.now()}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response



def exports_exam_marks_view(request):
    from main.models import Exam, MarksEntry, Grade

    if request.method == 'POST':
        grade_id = request.POST.get('grade_id')
        exam_id = request.POST.get('exam_id')
        # print(grade_id, exam_id)

        try:
            grade = Grade.objects.get(pk=grade_id)
            exam = Exam.objects.get(pk=exam_id)
        except (Grade.DoesNotExist, Exam.DoesNotExist):
            messages.error(request, 'Invalid grade or exam')
            return redirect('exports_exam_marks_view')

        marks_entries = MarksEntry.objects.filter(exam_paper__exam=exam, student__grade=grade)
        # print(marks_entries)
        if marks_entries.exists():
            temp_dict ={}
            th_plus_pr_subjects = set()
            for marks_entry in marks_entries:
                student_roll_no = marks_entry.student.roll_no
                subject = marks_entry.exam_paper.subject.name

                if student_roll_no not in temp_dict:
                    temp_dict[student_roll_no] = {}
                
                if subject not in temp_dict[student_roll_no]:
                    temp_dict[student_roll_no][subject] = {}
                    if marks_entry.exam_paper.practical_full_marks > 0:
                        th_plus_pr_subjects.add(subject)
                
                temp_dict[student_roll_no][subject] = {
                    'theory_marks': marks_entry.theory_marks,
                    'practical_marks': marks_entry.practical_marks
                }

        temp_dict = dict(sorted(temp_dict.items(), key=lambda x: x[0]))
        th_plus_pr_subjects = list(th_plus_pr_subjects)
        # print(th_plus_pr_subjects)
        import openpyxl
        from main.models import Student  # Student Model from main app
        from django.http import HttpResponse


        # Create a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Marks Data'
        

        FIRST_HEADER_CELLS = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1']
        SECOND_HEADER_CELLS = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2', 'P2', 'Q2', 'R2', 'S2', 'T2', 'U2', 'V2', 'W2', 'X2', 'Y2', 'Z2']
        #filling data into worksheet
        DATA_CELLS_ALPHA = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

        # Preparing Header for the worksheet
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws['A1'] = 'Roll No'
        ws['B1'] = 'Student Name'
        subject_current_cell_index = 2
        for subject in tuple(temp_dict.items())[0][1].keys():
            # print(subject, subject_current_cell_index)
            if subject in th_plus_pr_subjects:
                ws.merge_cells(f'{FIRST_HEADER_CELLS[subject_current_cell_index]}:{FIRST_HEADER_CELLS[subject_current_cell_index+1]}')
                ws[f'{FIRST_HEADER_CELLS[subject_current_cell_index]}'] = subject
                ws[f'{SECOND_HEADER_CELLS[subject_current_cell_index]}'] = 'PR'
                ws[f'{SECOND_HEADER_CELLS[subject_current_cell_index+1]}'] = 'TH'
                subject_current_cell_index += 2
            else:
                ws.merge_cells(f'{DATA_CELLS_ALPHA[subject_current_cell_index]}1:{DATA_CELLS_ALPHA[subject_current_cell_index]}2')
                ws[f'{FIRST_HEADER_CELLS[subject_current_cell_index]}'] = subject
                subject_current_cell_index += 1



        

        cell_index = 2
        row_index = 3
        for roll_no, data in temp_dict.items():
            student = Student.objects.get(grade=grade, roll_no=roll_no)
            ws[f'A{row_index}'] = roll_no
            ws[f'B{row_index}'] = student.name
            for subject, marks in data.items():
                # print(subject, marks)

                if subject in th_plus_pr_subjects:
                    ws[f'{DATA_CELLS_ALPHA[cell_index]}{row_index}'] = marks['practical_marks']
                    ws[f'{DATA_CELLS_ALPHA[cell_index+1]}{row_index}'] = marks['theory_marks']
                    cell_index += 2
                else:
                    ws[f'{DATA_CELLS_ALPHA[cell_index]}{row_index}'] = marks['theory_marks']
                    cell_index += 1

            cell_index = 2
            row_index += 1




        from .helpers import auto_fit_column_width
        auto_fit_column_width(ws)

        
        # Prepare the response as a downloadable Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        import datetime
        response['Content-Disposition'] = f'attachment; filename={exam}-students_marks_data-{datetime.datetime.now()}.xlsx'
        
        # Save the workbook to the response
        wb.save(response)
        
        return response

    else:
        context = {
            'exams': Exam.objects.all(),
            'grades': Grade.objects.all(),
        }

        return render(request, 'import_export/exam_marks_export.html', context)